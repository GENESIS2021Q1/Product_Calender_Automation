from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from datetime import datetime

gauth = GoogleAuth()
gauth.LoadCredentialsFile("credentials.txt")

drive = GoogleDrive(gauth)

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side



def upload_file(file_name):
    """Upload file -> change name in setcontentfile"""
    file1 = drive.CreateFile({"mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"})
    file1.SetContentFile(file_name)
    file1.Upload({"convert": True})


def download_file():
    """
    Searching for a file and downloading it -> insert file name after contains
    give file name to variable file_name
    """
    file_list = drive.ListFile({'q': "title contains 'FacultyCalendar_Output' and trashed=false"}).GetList()
    print(file_list[0]['title'])
    file_id = file_list[0]['id']
    print(file_id)
    file = drive.CreateFile({'id': file_id})
    file_time = datetime.now().strftime(" %Y-%m-%d_%I-%M-%S_%p")
    file_name = 'FacultyCalendar_Output'
    file_format = '.xlsx'
    file_title = file_name + file_time + file_format
    file.GetContentFile(file_title,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return file_title
def list_files():
    """
    Listing all files -> insert file name after contains
    This will list all the files that have similar name with their ID
    """
    file_list = drive.ListFile({'q': "title contains 'FacultyCalendar_Output' and trashed=false"}).GetList()
    for file in file_list:
        print(file['title'], file['id'])

"""Taking data from GUI"""

def FacultyCalendarFunction(SheetName, InputCalendar, Initiative, Month):
    #SheetName = "Sample_STEPin"
    #InputCalendar = "Automation_Sample Calender_v0.6.xlsx"
    #Initiative = "STEPin"
    #Month = "July"
    Month = Month.upper()
    wb = load_workbook(InputCalendar)
    InputExcel = wb[SheetName]
    all_rows = list(InputExcel.rows)
    all_Columns = list(InputExcel.columns)
    df = pd.read_excel('Automation_Sample Calender_v0.6.xlsx', 'Key')
    Initiatives = df['FixedInitiativeTitles']
    Initiativecode = df['FixedInitiativeCodes']
    Date = []
    FacultydList = []
    Leads1 = []
    Leads2 = []
    Leads3 = []
    SessionSlot = []
    RowLength = len(all_Columns[1])-3
    for head in range(8):
        for cell in range(3, len(all_Columns[head+1])):
            if head+1 == 1:
                list.append(Date, all_Columns[head+1][cell].value)
            elif head+1 == 2 | head+1 == 3 | head+1 == 4:
                break
            elif head+1 == 5:
                if all_Columns[head+1][cell].value:
                    temp = all_Columns[head+1][cell].value.strip()
                    temp.upper()
                    list.append(FacultydList, temp)
                    list.append(Leads1,temp)
                else:
                    list.append(Leads1, all_Columns[head+1][cell].value) 
            elif head+1 == 6:
                if all_Columns[head+1][cell].value:
                    temp = all_Columns[head+1][cell].value.strip()
                    temp.upper()
                    list.append(FacultydList, temp)
                    list.append(Leads2, temp)
                else:
                    list.append(Leads2, all_Columns[head+1][cell].value)                    
            elif head+1 == 7:
                if all_Columns[head+1][cell].value:
                    temp = all_Columns[head+1][cell].value.strip()
                    temp.upper()
                    list.append(FacultydList, temp)
                    list.append(Leads3, temp)
                else:
                    list.append(Leads3, all_Columns[head+1][cell].value) 
            elif head+1 == 8:
                if all_Columns[head+1][cell].value:
                    temp = all_Columns[head+1][cell].value.strip()
                    temp.upper()
                    list.append(SessionSlot, temp)
                else:
                    list.append(SessionSlot, all_Columns[head+1][cell].value) 
    FacultyList = []
    for val in FacultydList:
        if val != None :
            FacultyList.append(val)
    FacultyList = list(set(FacultyList))
    FacultyM1Slots = []
    FacultyM2Slots = []
    FacultyA1Slots = []
    FacultyA2Slots = []
    for i in range(len(FacultyList)):
        FacultyM1Slots.append([0]*31)
        FacultyM2Slots.append([0]*31)
        FacultyA1Slots.append([0]*31)
        FacultyA2Slots.append([0]*31)
    for k in range(len(Initiatives)):
        if Initiative == Initiatives[k]:
            code = k
    for i in range(RowLength):
        for j in range(len(FacultyList)):
            if Leads1[i] == FacultyList[j]:
                if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':
                    FacultyM1Slots[j][Date[i]-1] = str(Initiativecode[code])+'P'
                if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':
                    FacultyM2Slots[j][Date[i]-1] = str(Initiativecode[code])+'P'
                if SessionSlot[i] == 'A1' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                    FacultyA1Slots[j][Date[i]-1] = str(Initiativecode[code])+'P'
                if SessionSlot[i] == 'A2' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                    FacultyA2Slots[j][Date[i]-1] = str(Initiativecode[code])+'P'
            if Leads2[i] == FacultyList[j]:
                if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':                     
                    FacultyM1Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
                if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':
                    FacultyM2Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
                if SessionSlot[i] == 'A1' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                    FacultyA1Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
                if SessionSlot[i] == 'A2' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                    FacultyA2Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
            if Leads3[i] == FacultyList[j]:
                if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':                     
                    FacultyM1Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
                if SessionSlot[i] == 'M1' or SessionSlot[i] == 'M' or SessionSlot[i] == 'F':
                    FacultyM2Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
                if SessionSlot[i] == 'A1' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                    FacultyA1Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
                if SessionSlot[i] == 'A2' or SessionSlot[i] == 'A' or SessionSlot[i] == 'F':
                    FacultyA2Slots[j][Date[i]-1] = str(Initiativecode[code])+'S'
    Months = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER",
              "NOVEMBER", "DECEMBER"];
    for i in range(12):
        if Month == Months[i]:
            index = i
    Slots = []
    for i in range(len(FacultyList)):
        Slots.append([])
    for i in range(len(FacultyList)):
        for j in range(31):
            Slots[i].append(FacultyM1Slots[i][j])
            Slots[i].append(FacultyM2Slots[i][j])
            Slots[i].append(FacultyA1Slots[i][j])
            Slots[i].append(FacultyA2Slots[i][j])
    OutputExcelSheet = download_file()
    OutputSheet = pd.ExcelFile(OutputExcelSheet)
    OutputSheetNames = OutputSheet.sheet_names
    wb = load_workbook(OutputExcelSheet)
    OutputExcel = wb[OutputSheetNames[index]]
    OutputRows = list(OutputExcel.rows)
    ExdFacultyList = []
    for head in range(4,24):
        list.append(ExdFacultyList,OutputRows[head][1].value)    
    ExFacultyList = []
    for val in ExdFacultyList:
        if val != None :
            ExFacultyList.append(val)
    ExSlots = []
    for i in range(len(ExFacultyList)):
        ExSlots.append([])
    for head in range(4,4+len(ExFacultyList)):
        for cell in range(2,126):
            list.append(ExSlots[head-4],OutputRows[head][cell].value)
    CheckList = [0]*len(FacultyList)
    CheckList1 = [0]*len(ExFacultyList)
    Range = 0;
    ColourValues =['0000FF','FF0000','00FF00', '00800080', '00008080', '00FF99CC']
    for i in range(len(ExFacultyList)):
        for j in range(len(FacultyList)):
            if ExFacultyList[i] == FacultyList[j]:
                CheckList[j] = 1
                CheckList1[i] = 1
                for k in range(124):
                    if ExSlots[i][k] != Slots[j][k]:
                        if ExSlots[i][k]==0:
                            ExSlots[i][k] = Slots[j][k]
                        elif ExSlots[i][k] and Slots[j][k]!=0:
                            ExSlots[i][k] = (Slots[j][k]) + (ExSlots[i][k])
                        OutputExcel.cell(row=5+i, column=3+k).value = ExSlots[i][k]
                        if ExSlots[i][k]!=0:
                            temp = int(ExSlots[i][k][0])
                            OutputExcel.cell(row=5+i, column=3+k).fill = PatternFill(fill_type='solid', start_color=ColourValues[temp-1], end_color=ColourValues[temp-1])                            
    Range = 0;    
    for i in range(len(FacultyList)):
        if CheckList[i]==0:
            OutputExcel.cell(row=5+len(ExFacultyList)+Range, column=2).value = FacultyList[i]
            for j in range(124):
                OutputExcel.cell(row=5+len(ExFacultyList)+Range, column=3+j).value = Slots[i][j]
                if Slots[i][j]!=0:
                    temp = int(Slots[i][j][0])
                    OutputExcel.cell(row=5+len(ExFacultyList)+Range, column=3+j).fill = PatternFill(fill_type='solid', start_color=ColourValues[temp-1], end_color=ColourValues[temp-1])
            Range = Range + 1
    wb.save(OutputExcelSheet)
    upload_file(OutputExcelSheet)


if __name__ == '__main__':
    SheetName = "Sample_STEPin"
    InputCalendar = "Automation_Sample Calender_v0.6.xlsx"
    Initiative = "STEPin"
    Month = "July"
    FacultyCalendarFunction(SheetName, InputCalendar, Initiative, Month)

